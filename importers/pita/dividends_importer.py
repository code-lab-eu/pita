from dividend_payment import DividendPayment
from openpyxl import load_workbook
import decimal
from dividend_collection import DividendCollection
from typeguard import typechecked

COLUMNS = [
    "dividend_date",
    "security",
    "company",
    "country",
    "dividend",
    "tax_paid",
]


class PitaDividendsImporter:
    @staticmethod
    @typechecked
    def import_dividends(collection: DividendCollection, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        # Skip the first row, because it contains column headers.
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip the row if it is empty.
            if not row[COLUMNS.index("dividend_date")]:
                continue

            dividend_date = row[COLUMNS.index("dividend_date")]
            security = row[COLUMNS.index("security")]
            company = row[COLUMNS.index("company")]
            country = row[COLUMNS.index("country")]
            dividend = decimal.Decimal(row[COLUMNS.index("dividend")])
            tax_paid = decimal.Decimal(row[COLUMNS.index("tax_paid")])

            dividend_payment = DividendPayment(
                date_time=dividend_date,
                fund_name=security,
                company=company,
                country=country,
                dividend=dividend,
                tax_paid=tax_paid,
                currency="EUR",
                purchase_price=0.00,
            )
            collection.append(dividend_payment)
