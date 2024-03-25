from openpyxl import load_workbook
from transaction import Transaction
import decimal

COLUMNS = [
    "fund_name",
    "domicile",
    "date_time",
    "transaction_type",
    "number",
    "share_price",
    "total_shares",
    "purchase_price",
    "profit_loss",
]


class PitaTransactionsImporter:
    @staticmethod
    def import_transactions(collection, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        # Skip the first row, because it contains column headers.
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip the row if the fund name is missing.
            if not row[COLUMNS.index("fund_name")]:
                continue

            fund_name = row[COLUMNS.index("fund_name")]
            domicile = row[COLUMNS.index("domicile")]
            date_time = row[COLUMNS.index("date_time")]
            transaction_type = row[COLUMNS.index("transaction_type")]
            number = decimal.Decimal(row[COLUMNS.index("number")])
            share_price = decimal.Decimal(row[COLUMNS.index("share_price")])
            purchase_price = decimal.Decimal(row[COLUMNS.index("purchase_price")])
            profit_loss = row[COLUMNS.index("profit_loss")]
            profit_loss = None if profit_loss is None else decimal.Decimal(row[COLUMNS.index("profit_loss")])
            transaction = Transaction(
                fund_name=fund_name,
                domicile=domicile,
                date_time=date_time,
                transaction_type=transaction_type,
                number=number,
                share_price=share_price,
                purchase_price=purchase_price,
                profit_loss=profit_loss,
                currency="EUR",
            )
            collection.append(transaction)
