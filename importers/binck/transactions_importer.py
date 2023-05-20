from transaction import Transaction
from openpyxl import load_workbook
import decimal


class BinckTransactionsImporter:
    @staticmethod
    def import_transactions(collection, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        # We are starting from row 2, because the first row is a title.
        for i in range(2, sheet.max_row + 1):
            # If we don't have the fund this row is empty. Skip it!
            fund_name = sheet.cell(row=i, column=1).value
            if not fund_name:
                continue
            date_time = sheet.cell(row=i, column=3).value
            domicile = sheet.cell(row=i, column=2).value
            transaction_type = sheet.cell(row=i, column=4).value
            number = decimal.Decimal(sheet.cell(row=i, column=5).value)
            share_price = decimal.Decimal(sheet.cell(row=i, column=6).value)
            currency = "EUR"
            purchase_price = share_price * number
            transaction = Transaction(
                fund_name,
                domicile,
                date_time,
                transaction_type,
                number,
                share_price,
                currency,
                purchase_price,
            )
            collection.append(transaction)
