from dividend_payment import DividendPayment
from dividend_collection import DividendCollection
from openpyxl import load_workbook
import decimal
import datetime
from typeguard import typechecked

FUND_DOMICILE_MAPPING = {
    "Amundi Index FTSE EPRA Nareit Global- ETF": "Luxembourg",
    "iShares Core Euro Corporate Bond UCITS ETF": "Ireland",
    "iShares Core Global Aggregate Bond UCITS ETF": "Ireland",
    "iShares Core MSCI World UCITS ETF": "Ireland",
    "iShares EM Infrastructure UCITS ETF": "Ireland",
    "iShares Emerging Markets Local Gov Bond UCITS ETF": "Ireland",
    "iShares FTSE/EPRA European Property EUR UCITS ETF": "Ireland",
    "iShares High Yield Corp. Bond UCITS ETF": "Ireland",
    "iShares MSCI Turkey UCITS ETF": "Ireland",
    "iShares MSCI World Small Cap UCITS ETF": "Ireland",
    "iShares S&P Global Clean Energy ETF": "Ireland",
    "iShares USD High Yield Capped Bond UCITS ETF": "Ireland",
}


class SaxoDividendsImporter:
    @staticmethod
    @typechecked
    def import_dividends(collection: DividendCollection, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        # Check that we have the required headers, and create a mapping from header name to column index.
        required_headers = [
            "Instrument",
            "Booked Amount (EUR)",
            "Type",
            "Account Currency",
            "Total Tax (EUR)",
            "Pay Date",
        ]
        headers = {}

        for required_header in required_headers:
            header_found = False
            for i in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=i).value
                if header == required_header:
                    headers[required_header] = i
                    header_found = True
                    break
            if not header_found:
                print(
                    "Error: Could not find required header "
                    + required_header
                    + " in file "
                    + file_name
                )
                exit()

        # We are starting from row 2, because the first row is a title.
        for i in range(2, sheet.max_row + 1):
            # If we don't have the fund this row is empty. Skip it!
            security = sheet.cell(row=i, column=headers["Instrument"]).value
            if not security:
                continue
            if sheet.cell(row=i, column=headers["Type"]).value != "Cash Dividend":
                continue

            date_time = sheet.cell(row=i, column=headers["Pay Date"]).value
            # Convert the date in format "'17-May-2021" to a datetime object.
            date_time = datetime.datetime.strptime(date_time, "%d-%b-%Y")
            company = security.split(" ")[0]
            dividend = decimal.Decimal(
                sheet.cell(row=i, column=headers["Booked Amount (EUR)"]).value
            )
            currency = sheet.cell(row=i, column=headers["Account Currency"]).value
            tax_paid = sheet.cell(row=i, column=headers["Total Tax (EUR)"]).value
            country = FUND_DOMICILE_MAPPING.get(security)
            if not country:
                print(
                    "Error could not find domicile "
                    + security
                    + " add it to FUND_DOMICILE_MAPPING!"
                )
                exit()
            dividend_payment = DividendPayment(
                security,
                company,
                dividend,
                country,
                tax_paid,
                currency,
                date_time,
            )
            collection.append(dividend_payment)
