from transaction import Transaction
from openpyxl import load_workbook
import decimal
import datetime

FUND_DOMICILE_MAPPING = {
    "Amundi Index FTSE EPRA Nareit Global- ETF": "Luxembourg",
    "iShares Core Euro Corporate Bond UCITS ETF": "Ireland",
    "iShares Core Global Aggregate Bond UCITS ETF": "Ireland",
    "iShares Core MSCI World UCITS ETF": "Ireland",
    "iShares EM Infrastructure UCITS ETF" : "Ireland",
    "iShares Emerging Markets Local Gov Bond UCITS ETF": "Ireland",
    "iShares FTSE/EPRA European Property EUR UCITS ETF": "Ireland",
    "iShares High Yield Corp. Bond UCITS ETF": "Ireland",
    "iShares MSCI Turkey UCITS ETF": "Ireland",
    "iShares MSCI World Small Cap UCITS ETF": "Ireland",
    "iShares S&P Global Clean Energy ETF": "Ireland",
    "iShares USD High Yield Capped Bond UCITS ETF": "Ireland",
}


class SaxoImporter:
    @staticmethod
    def import_transactions(collection, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active
        sheet2 = workbook.worksheets[1]

        # Check that we have the required headers, and create a mapping from header name to column index.
        required_headers = ["Instrument", "TradeTime", "B/S", "Amount", "Trade Value", "Price", "Account ID", "Open/Close"]
        headers = {}

        for required_header in required_headers:
            header_found = False
            for i in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=i).value
                if header == required_header:
                    headers[required_header] = i
                    header_found = True
                    break
            if not header_found:
                print("Error: Could not find required header " + required_header + " in file " + file_name)
                exit()

        # We are starting from row 2, because the first row is a title.
        for i in range(2, sheet.max_row + 1):

            # Skip sales, these are handled by import_closed_positions()
            if sheet.cell(row=i, column=headers["Open/Close"]).value != "Open":
                continue

            date_time = sheet.cell(row=i, column=headers['TradeTime']).value
            # Convert the date in format "'17-May-2021" to a datetime object.
            date_time = datetime.datetime.strptime(date_time, "%d-%b-%Y")

            # The most clean fund name is in the second sheet.
            fund_name = sheet2.cell(row=i, column=42).value
            domicile = FUND_DOMICILE_MAPPING.get(fund_name)
            transaction_type = sheet.cell(row=i, column=headers['B/S']).value
            number = decimal.Decimal(sheet.cell(row=i, column=headers["Amount"]).value)
            share_price = decimal.Decimal(sheet.cell(row=i, column=headers["Price"]).value)
            currency = sheet.cell(row=i, column=headers["Account ID"]).value[-3:]
            purchase_price = share_price * number
            transaction = Transaction(fund_name, domicile, date_time, transaction_type, number, share_price,
                                      currency, purchase_price)
            collection.append(transaction)

    @staticmethod
    def import_closed_positions(collection, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        # Check that we have the required headers, and create a mapping from header name to column index.
        required_headers = ["Instrument Description", "Trade Date Close", "Account Currency", "AmountClose", "Close Price", "PnLAccountCurrency"]
        headers = {}

        for required_header in required_headers:
            header_found = False
            for i in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=i).value
                if header == required_header:
                    headers[required_header] = i
                    header_found = True
                    break
            if not header_found:
                print("Error: Could not find required header " + required_header + " in file " + file_name)
                exit()

        # We are starting from row 2, because the first row is a title.
        for i in range(2, sheet.max_row + 1):

            fund_name = sheet.cell(row=i, column=headers["Instrument Description"]).value

            date_time = sheet.cell(row=i, column=headers['Trade Date Close']).value
            # Convert the date in format "'17-May-2021" to a datetime object.
            date_time = datetime.datetime.strptime(date_time, "%d-%b-%Y")

            domicile = FUND_DOMICILE_MAPPING.get(fund_name)
            transaction_type = "Sold"
            number = decimal.Decimal(sheet.cell(row=i, column=headers["AmountClose"]).value)
            share_price = decimal.Decimal(sheet.cell(row=i, column=headers["Close Price"]).value)
            currency = sheet.cell(row=i, column=headers["Account Currency"]).value[-3:]
            purchase_price = share_price * number
            profit_loss = decimal.Decimal(sheet.cell(row=i, column=headers["PnLAccountCurrency"]).value)
            transaction = Transaction(fund_name, domicile, date_time, transaction_type, number, share_price,
                                      currency, purchase_price, profit_loss)
            collection.append(transaction)
