import argparse
import datetime
import sys

from importers.binck.transactions_importer import BinckTransactionsImporter
from importers.pita.dividends_importer import PitaDividendsImporter
from importers.pita.transactions_importer import PitaTransactionsImporter
from importers.trading212.dividends_importer import Trading212DividendsImporter
from importers.trading212.transactions_importer import Trading212TransactionsImporter
from importers.saxo.dividends_importer import SaxoDividendsImporter
from importers.saxo.transactions_importer import SaxoImporter

from transaction_collection import TransactionCollection
from dividend_collection import DividendCollection
from appendix8row import Appendix8Row

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--binck-transactions",
        "--binck",
        dest="transactions_binck",
        help="The path to an Excel file detailing BinckBank transactions.",
    )
    parser.add_argument(
        "--trading212-transactions",
        "--t212",
        dest="transactions_trading212",
        help="The path to an csv file detailing Trading 212 transactions + dividends.",
        action="append",
        nargs="*",
    )
    parser.add_argument(
        "--saxo-trades",
        dest="saxo_trades",
        help="The path to an Excel file detailing Saxo trades.",
    )
    parser.add_argument(
        "--saxo-dividends",
        dest="saxo_dividends",
        help="The path to an Excel file detailing Saxo dividends.",
    )
    parser.add_argument(
        "--saxo-closed-positions",
        dest="saxo_closed_positions",
        help="The path to an Excel file detailing Saxo closed positions.",
    )
    parser.add_argument(
        "--pita-investments",
        dest="pita_investments",
        help="The path to an Excel file containing our own exported investments.",
    )
    parser.add_argument(
        "--pita-dividends",
        dest="pita_dividends",
        help="The path to an Excel file containing our own exported dividends.",
    )
    args = parser.parse_args()

    # If the user doesn't supply input print the help and exit.
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(1)

    transactions = TransactionCollection()
    payments = DividendCollection()

    if args.transactions_trading212:
        # Loop over the CSV files.
        for csv_files in args.transactions_trading212:
            # Argparse allows to add multiple values for an option in a row.
            for csv_file in csv_files:
                Trading212TransactionsImporter.import_transactions(
                    transactions, csv_file
                )
                Trading212DividendsImporter.import_dividends(payments, csv_file)

    if args.transactions_binck:
        BinckTransactionsImporter.import_transactions(
            transactions, args.transactions_binck
        )

    if args.saxo_trades:
        SaxoImporter.import_transactions(transactions, args.saxo_trades)

    if args.saxo_dividends:
        SaxoDividendsImporter.import_dividends(payments, args.saxo_dividends)

    if args.saxo_closed_positions:
        SaxoImporter.import_closed_positions(transactions, args.saxo_closed_positions)

    if args.pita_investments:
        PitaTransactionsImporter.import_transactions(transactions, args.pita_investments)

    if args.pita_dividends:
        PitaDividendsImporter.import_dividends(payments, args.pita_dividends)

    # Export the transactions to an Excel file.
    if not transactions.is_empty():
        wb = Workbook()

        # Grab the active worksheet.
        ws = wb.active

        # Create the header row.
        ws["A1"] = "Fund name"
        ws["B1"] = "Domicile"
        ws["C1"] = "Transaction date"
        ws["D1"] = "Transaction type"
        ws["E1"] = "Number"
        ws["F1"] = "Share price"
        ws["G1"] = "Total shares"
        ws["H1"] = "Purchase price"
        ws["I1"] = "Profit/Loss"

        # Set a blue background color and white font color on row 1 (the header row).
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="0066CC", fill_type="solid")
            cell.font = Font(name="Calibri", color="FFFFFF")

        # Column widths.
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15

        transactions.calc_total_shares()
        transactions.sort_by_date(True)
        fund_names = transactions.get_fund_names()
        for fund_name in fund_names:
            fund_transactions = transactions.get_fund_transactions(fund_name)

            # Add each transaction as a new row in the Excel sheet.
            for transaction in fund_transactions:
                ws.append(
                    [
                        transaction.fund_name,
                        transaction.domicile,
                        transaction.date_time,
                        transaction.transaction_type,
                        transaction.number,
                        transaction.share_price,
                        transaction.total,
                        transaction.purchase_price,
                        transaction.profit_loss,
                    ]
                )

                # Format the date as DD.MM.YYYY.
                ws.cell(ws.max_row, 3).number_format = "DD.MM.YYYY"

                # Format negative number in number column.
                number_format = "#,0.000;[RED]-#,0.000"
                # Format and styling the negative number in number column.
                ws.cell(ws.max_row, 5).number_format = number_format
                # Format and styling total share column.
                ws.cell(ws.max_row, 7).number_format = number_format

                # Format the currencies.
                # See https://support.microsoft.com/en-us/office/number-format-codes-5026bbd6-04bc-48cd-bf33-80f18b4eae68?ui=en-us&rs=en-us&ad=us
                number_format = (
                    "#,##0.00 [$"
                    + transaction.currency
                    + "];[RED]-#,##0.00 [$"
                    + transaction.currency
                    + "]"
                )
                # Format the share price as a currency.
                ws.cell(ws.max_row, 6).number_format = number_format
                # Format the purchase price as a currency.
                ws.cell(ws.max_row, 8).number_format = number_format
                # Format the profit/loss as a currency if it is not empty.
                if transaction.profit_loss != "":
                    ws.cell(ws.max_row, 9).number_format = number_format

            # Insert an empty row inbetween each fund.
            ws.append([])

        # Save the file
        wb.save("investments.xlsx")
        print("Saved file investments.xlsx!")

        # Sales report
        if not transactions.is_empty():
            wb = Workbook()

            # Grab the active worksheet.
            ws = wb.active

            # Create the header row.
            ws["A1"] = "Security"
            ws["B1"] = "Company"
            ws["C1"] = "Country"
            ws["D1"] = "Profit"

            # Set a blue background color and white font color on row 1 (the header row).
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="0066CC", fill_type="solid")
                cell.font = Font(name="Calibri", color="FFFFFF")

            # Column widths.
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 20

            fund_names = transactions.get_fund_names()
            for fund_name in fund_names:
                fund_transactions = transactions.get_fund_transactions(fund_name)

                profit = 0
                transaction = None
                for transaction in fund_transactions:
                    # We only need to report sales in the past year.
                    if transaction.transaction_type != "Sell":
                        continue
                    if transaction.date_time.year != datetime.date.today().year - 1:
                        continue
                    profit += transaction.profit_loss

                if transaction:
                    fund_name = transaction.fund_name
                    company = transaction.fund_name.split(" ")[0]
                    ws.append(
                        [transaction.fund_name, company, transaction.domicile, profit]
                    )

                    # Format the currency.
                    number_format = (
                        "#,##0.00 [$"
                        + transaction.currency
                        + "];[RED]-#,##0.00 [$"
                        + transaction.currency
                        + "]"
                    )
                    ws.cell(ws.max_row, 4).number_format = number_format

            # Save the file
            wb.save("sales.xlsx")
            print("Saved file sales.xlsx!")

        # Dividend payments
        if not payments.is_empty():
            wb = Workbook()

            # grab the active worksheet
            ws = wb.active

            # Create the header row.
            ws["A1"] = "Dividend date"
            ws["B1"] = "Security"
            ws["C1"] = "Company"
            ws["D1"] = "Country"
            ws["E1"] = "Dividend"
            ws["F1"] = "Tax paid/withheld abroad"

            # Set a blue background color and white font color on row 1 (the header row).
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="0066CC", fill_type="solid")
                cell.font = Font(name="Calibri", color="FFFFFF")

            # Column width
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 30

            fund_names = payments.get_fund_names()
            for fund_name in fund_names:
                fund_payments = payments.get_fund_payments(fund_name)

                # Add each transaction as a new row in the Excel sheet.
                for payment in fund_payments:
                    # We only need to report the dividends for last year.
                    last_year = datetime.date.today().year - 1
                    if payment.date_time.year != last_year:
                        continue
                    ws.append(
                        [
                            payment.date_time,
                            payment.fund_name,
                            payment.company,
                            payment.country,
                            payment.dividend,
                            payment.tax_paid,
                        ]
                    )

                    # Format the date as DD.MM.YYYY.
                    ws.cell(ws.max_row, 1).number_format = "DD.MM.YYYY"
                    # Format the currencies.
                    # See https://support.microsoft.com/en-us/office/number-format-codes-5026bbd6-04bc-48cd-bf33-80f18b4eae68?ui=en-us&rs=en-us&ad=us
                    number_format = (
                        "#,##0.00 [$"
                        + payment.currency
                        + "];[RED]-#,##0.00 [$"
                        + payment.currency
                        + "]"
                    )
                    # Format the share price as a currency.
                    ws.cell(ws.max_row, 5).number_format = number_format
                    # Format the purchase price as a currency.
                    ws.cell(ws.max_row, 6).number_format = number_format

                # Insert an empty row inbetween each fund.
                ws.append([])

            # Save the file
            wb.save("dividends.xlsx")
            print("Saved file dividends.xlsx!")

            # Export the Excel sheet for appendix 8.
            if not transactions.is_empty():
                wb = Workbook()

                # Grab the active worksheet.
                ws = wb.active

                # Create the header row.
                ws["A1"] = "Domicile"
                ws["B1"] = "Number"
                ws["C1"] = "Transaction date"
                ws["D1"] = "Total price origin currency"
                ws["E1"] = "Total price in BGN"

                # Set a blue background color and white font color on row 1 (the header row).
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="0066CC", fill_type="solid")
                    cell.font = Font(name="Calibri", color="FFFFFF")

                # Column widths.
                ws.column_dimensions["A"].width = 20
                ws.column_dimensions["B"].width = 20
                ws.column_dimensions["C"].width = 20
                ws.column_dimensions["D"].width = 20
                ws.column_dimensions["E"].width = 20

                transactions.sort_by_date(False)

                # For the Appendix 8 report we need a list of the number of stocks that have been bought on a particular
                # day, along with the price in the original currency and in leva.
                appendix8rows = {}

                for transaction in transactions.transactions:
                    date = transaction.date_time.date()
                    domicile = transaction.domicile
                    # Check if we already have a row for the current date and domicile. If not, create one.
                    if not (date, domicile) in appendix8rows:
                        appendix8rows[(date, domicile)] = Appendix8Row(date, domicile)
                    appendix8rows[(date, domicile)].add_transaction(transaction)

                # Add the rows to the Excel sheet.
                for row in appendix8rows.values():
                    ws.append(
                        [
                            row.domicile,
                            row.number,
                            row.date,
                            row.price_eur,
                            row.price_eur * Decimal(1.95583),
                        ]
                    )

                    # Format the date as DD.MM.YYYY.
                    ws.cell(ws.max_row, 3).number_format = "DD.MM.YYYY"

                    # Format the currencies.
                    # See https://support.microsoft.com/en-us/office/number-format-codes-5026bbd6-04bc-48cd-bf33-80f18b4eae68?ui=en-us&rs=en-us&ad=us
                    number_format_eur = "#,##0.00 [$EUR];[RED]-#,##0.00 [$EUR]"
                    number_format_bgn = "#,##0.00 [$BGN];[RED]-#,##0.00 [$BGN]"
                    # Format the share price as a currency.
                    ws.cell(ws.max_row, 4).number_format = number_format_eur
                    # Format the purchase price as a currency.
                    ws.cell(ws.max_row, 5).number_format = number_format_bgn

                # Save the file
                wb.save("investments_appendix8.xlsx")
                print("Saved file investments_appendix8.xlsx!")
